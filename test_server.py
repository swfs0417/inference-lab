import io
import os
import tempfile
import unittest
from pathlib import Path

import server


class MetricsTests(unittest.TestCase):
    def test_percentile_interpolates(self):
        self.assertEqual(server.percentile([100, 200, 300, 400], .50), 250)

    def test_summary_ignores_failed_samples(self):
        summary = server.summarize([
            {"ttft_ms": 100, "total_ms": 500, "output_tokens_per_sec": 20, "error": None},
            {"ttft_ms": 200, "total_ms": 700, "output_tokens_per_sec": 30, "error": None},
            {"total_ms": 50, "error": "timeout"},
        ])
        self.assertEqual(summary["requests"], 3)
        self.assertEqual(summary["successes"], 2)
        self.assertEqual(summary["errors"], 1)
        self.assertEqual(summary["ttft_mean_ms"], 150)
        self.assertEqual(summary["output_tps_mean"], 25)

    def test_itl_summary(self):
        summary = server.summarize([
            {"ttft_ms": 50, "total_ms": 100, "output_tokens_per_sec": 40,
             "prompt_tokens": 10, "output_tokens": 4, "inter_chunk_ms": [8, 10, 12], "error": None}
        ])
        self.assertEqual(summary["itl_mean_ms"], 10)
        self.assertEqual(summary["itl_p95_ms"], 11.8)
        self.assertEqual(summary["prompt_tokens_mean"], 10)

    def test_gpu_summary_and_energy(self):
        summary = server.summarize_gpu([
            {"at_ms": 0, "gpus": [{"gpu_util_pct": 20, "memory_used_mb": 1000, "power_w": 100, "temperature_c": 50}]},
            {"at_ms": 3600, "gpus": [{"gpu_util_pct": 80, "memory_used_mb": 2000, "power_w": 100, "temperature_c": 60}]},
        ])
        self.assertEqual(summary["gpu_util_mean_pct"], 50)
        self.assertEqual(summary["memory_peak_mb"], 2000)
        self.assertEqual(summary["energy_wh"], 0.1)

    def test_endpoint_allowlist(self):
        previous = os.environ.get("ALLOWED_TARGET_HOSTS")
        os.environ["ALLOWED_TARGET_HOSTS"] = "localhost,10.1.2.3"
        try:
            server.validate_endpoint("http://10.1.2.3:8000/v1")
            with self.assertRaises(ValueError):
                server.validate_endpoint("https://example.com/v1")
        finally:
            if previous is None:
                os.environ.pop("ALLOWED_TARGET_HOSTS", None)
            else:
                os.environ["ALLOWED_TARGET_HOSTS"] = previous

    def test_rejects_invalid_reasoning_effort(self):
        with self.assertRaisesRegex(ValueError, "reasoning_effort"):
            server.run_benchmark({
                "endpoint": "http://127.0.0.1:11434/v1", "model": "test",
                "prompt": "test", "reasoning_effort": "maximum",
            })


class PersistenceTests(unittest.TestCase):
    def test_round_trip(self):
        previous = server.DB_PATH
        try:
            with tempfile.TemporaryDirectory() as tmp:
                server.DB_PATH = Path(tmp) / "test.db"
                server.init_db()
                server.save_run({
                    "id": "test-run", "source": "test", "model": "model",
                    "endpoint": "local", "settings": {}, "summary": {"requests": 1},
                    "samples": [{"total_ms": 1, "output_text": "저장된 응답"}], "gpu": [],
                })
                run = server.list_runs(1)[0]
                self.assertEqual(run["id"], "test-run")
                self.assertEqual(run["summary"]["requests"], 1)
                self.assertEqual(run["samples"][0]["output_text"], "저장된 응답")
        finally:
            server.DB_PATH = previous


class ExportTests(unittest.TestCase):
    def setUp(self):
        self.runs = [{
            "id": "run-1", "created_at": "2026-08-08 12:00:00", "source": "test",
            "model": "model", "endpoint": "local", "settings": {"prompt": "=unsafe", "condition": "warm"},
            "summary": {"requests": 1, "successes": 1, "errors": 0, "ttft_mean_ms": 100},
            "samples": [{
                "sequence": 1, "phase": "first-measured", "ttft_ms": 100,
                "total_ms": 500, "output_text": "=output",
            }],
        }]

    def test_csv_export_includes_prompt_and_blocks_formula_injection(self):
        csv_text = server.export_runs_csv(self.runs).decode("utf-8-sig")
        self.assertIn("prompt", csv_text)
        self.assertIn("'=unsafe", csv_text)
        self.assertIn("'=output", csv_text)

    def test_xlsx_export_has_run_and_sample_sheets(self):
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(server.export_runs_xlsx(self.runs)), read_only=True)
        self.assertEqual(workbook.sheetnames, ["Runs", "Samples"])
        self.assertEqual(workbook["Runs"]["F2"].value, "'=unsafe")
        self.assertEqual(workbook["Samples"]["A2"].value, "run-1")
        self.assertEqual(workbook["Samples"]["P2"].value, "'=output")

    def test_pdf_converter_rejects_non_pdf(self):
        with self.assertRaisesRegex(ValueError, "PDF 파일만"):
            server.convert_pdf_to_markdown("note.txt", "aGVsbG8=")


if __name__ == "__main__":
    unittest.main()
